import streamlit as st
st.set_page_config(page_title="Análisis Granulométrico", layout="wide")
import pandas as pd
import numpy as np

#from scipy.interpolate import PchipInterpolator
from io import BytesIO
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap
import matplotlib.pyplot as plt
#import functions as fn
from functools import reduce
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
import sys
import os
from cryptography.fernet import Fernet
import streamlit_authenticator as stauth
import yaml
from yaml.loader import SafeLoader

from streamlit_plugins.components.theme_changer import get_active_theme_key, st_theme_changer
from streamlit_plugins.components.theme_changer.entity import ThemeInfo, ThemeInput, ThemeBaseLight, ThemeBaseDark
from streamlit_plugins.components.theme_changer import get_active_theme_key



with open('config.yaml') as file:
    config = yaml.load(file, Loader=SafeLoader)

authenticator = stauth.Authenticate(
    config['credentials'],
    config['cookie']['name'],
    config['cookie']['key'],
    config['cookie']['expiry_days']
)

name, auth_status, username = authenticator.login('Login', 'main')
# 4. Gestión de estados de autenticación
if auth_status:
    st.sidebar.success(f"Welcome *{name}*")
    # Logout en sidebar
    authenticator.logout('Logout', 'sidebar')
    # Aquí va el resto de tu app protegida
    st.write("🔒 Aplicación protegida")
    # … procesamiento, gráficas, descargas, etc.

elif auth_status is False:
    st.error("❌ Username/password incorrect")

else:  # auth_status is None
    st.info("ℹ️ Please enter your credentials or use guest login")
    # Solo si aún no está logueado mostramos opciones de invitado
    authenticator.experimental_guest_login(
        'Login with Google',
        provider='google',
        oauth2=config['oauth2']
    )
    authenticator.experimental_guest_login(
        'Login with Microsoft',
        provider='microsoft',
        oauth2=config['oauth2']
    )

# Obtener la clave desde GitHub Actions (ya configurada como secret)

key_str = st.secrets["STREAMLIT_GRANULOMETRIA_KEY"]

if not key_str:
    raise RuntimeError("No se encontró la clave STREAMLIT_GRANULOMETRIA_KEY")
key = key_str.encode()

fernet = Fernet(key)

# Leer y descifrar functions.py.enc
with open("functions.py.enc", "rb") as f:
    decrypted_code = fernet.decrypt(f.read())

# Crear un namespace aislado para las funciones
fn = {}

# Ejecutar el código descifrado dentro del dict `fn`
exec(decrypted_code, fn)
# print("Test")
# fn2["prueba_encriptado"]("Fran")


try:
    if '_pydevd_frame_eval.pydevd_frame_eval_cython_wrapper' not in sys.modules:
        import _pydevd_frame_eval.pydevd_frame_eval_cython_wrapper
except ImportError:
    pass





#active_theme = get_active_theme_key()
# Inyectar fuente Roboto

st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Roboto:wght@400;700&display=swap');

    /* Aplicar Roboto */
    h1, h2, h3, h4, h5, h6,
    .stMarkdown, .stText, .stTitle, .stHeader {
        font-family: 'Roboto',sans-serif !important;
    }

    # /* Sidebar y fondo */
    # section[data-testid="stSidebar"] {
    #     background-color: #002f42;
    # }

    div[data-testid="stAppViewContainer"] > main {
        background-color: #002f42;
    }

    body {
        background-color: #002f42;
    }

    /* Uploader */
    div[data-testid="stFileUploader"] {
        background-color: #002f42;
        padding: 1rem;
        border-radius: 0.5rem;
        border: 1px solid #002f42;
    }

    /*  Cambiar color del texto del uploader */
    section[data-testid="stSidebar"] label {
        color: white !important;  /* Cambialo por el color que quieras */
        font-family: 'Roboto', sans-serif;
    }

    /* Footer */
    .footer {
        position: fixed;
        bottom: 0;
        left: 0;
        width: 100%;
        background-color: #101820;
        color: white;
        display: flex;
        justify-content: center; /* Centra el contenido principal */
        align-items: center;
        padding: 10px 60px 10px 20px;  /* espacio a la derecha para el logo */
        font-size: 14px;
        z-index: 100;
        box-sizing: border-box;
    }

    .footer-content {
        text-align: center;
        flex-grow: 1;
    }

    .footer img {
        height: 30px;
        margin-left: auto;
        margin-right: 120px; /* ajusta este valor según cuán a la derecha lo quieras */
    }

    .footer a {
        color: #ffffff;
        text-decoration: none;
        font-weight: bold;
    }

    .logo-container {
        position: absolute;
        right: 0;
        padding-right: 20px;
    }
    </style>


    <div class="footer">
        <div class="footer-content">
            <a href="mailto:francisco.corvalan6@gmail.com">📧 Contacto: francisco.corvalan6@gmail.com</a>
        </div>

    </div>
    """, unsafe_allow_html=True)

######################################################################################################

init_theme_data = dict(
    soft_light=ThemeInput(
        name="Soft Sunrise",
        icon=":material/sunny_snowing:",
        order=0,
        themeInfo=ThemeInfo(
            base=ThemeBaseLight.base,
            primaryColor="#002f42",
            backgroundColor="#d1dde6",
            secondaryBackgroundColor="#01506e",
            textColor="#101820",
            widgetBackgroundColor="#F3F5F7",
            widgetBorderColor="#101820",
            skeletonBackgroundColor="#f0f8ff",
            #bodyFont=ThemeBaseLight.bodyFont,
            #codeFont=ThemeBaseLight.codeFont,
            #fontFaces=ThemeBaseLight.fontFaces,
        )
    ),
    soft_dark=ThemeInput(
        name="Dark Midnight",
        icon=":material/nights_stay:",
        order=1,
        themeInfo=ThemeInfo(
            base=ThemeBaseDark.base,
            primaryColor="#0095c8", # Color pestaña seleccionada
            backgroundColor="#000000",
            secondaryBackgroundColor="#002f42",
            textColor="#f0f8ff",
            widgetBackgroundColor="#002f42",
            widgetBorderColor="#002f42",
            skeletonBackgroundColor="#f0f8ff",
            #bodyFont=ThemeBaseDark.bodyFont,
            #codeFont=ThemeBaseDark.codeFont,
            #fontFaces=ThemeBaseDark.fontFaces,
        )
    )
)
if st.session_state.get("theme_data") is None:
    st.session_state["theme_data"] = init_theme_data

theme_data = st.session_state["theme_data"]

st_theme_changer(themes_data=theme_data, render_mode="init", default_init_theme_name="soft_dark")
#st_theme_changer(themes_data=theme_data, rerun_whole_st=True)


df_transformado = None

 
st.logo(
    "./Logo/ausenco-logo.png",
    #link="https://www.ausenco.com/",
    size="large"
    #icon_image="./Logo/ausenco-logo.png",
)
tabs = st.tabs(["Aplicación", "Manual de Usuario",' Metodología'])

with tabs[0]:
    # Configuración de la página
    
    # Título de la aplicación
    st.title("📊 Análisis de Curvas Granulométricas")
    st.markdown("""
    Suba su archivo Excel con datos granulométricos y personalice la visualización.
    La aplicación graficará automáticamente las curvas en escala logarítmica.
    """)
    # Barra lateral para carga de archivos
    with st.sidebar:
        #st.header("Carga y Configuración")
        archivo_subido = st.file_uploader("Subir archivo Excel", type=['xlsx', 'xls'],
                                        help="Arrastra y suelta tu archivo o haz clic para seleccionarlo"  # Texto de ayuda
                                        )
        print('archivo_subido',archivo_subido)
        if archivo_subido is not None:
            try:
                df = pd.read_excel(archivo_subido)
                df_transformado,df_plasticidad = fn["df_process"](df) # fn.df_process(df)
                # Asumimos que la primera columna es 'Tamaño' y las demás son muestras
                columna_tamaño = 'Tamaño'
                columnas_muestras = df.columns[1:]
                df_transformado = df_transformado.dropna()
                
                # Obtenemos el rango completo de tamaños (en escala log10)
                min_log = df_transformado['Tamaño'].min()
                max_log = df_transformado['Tamaño'].max()
                
                st.success("¡Archivo cargado correctamente!")

                # Opción para mostrar puntos
                mostrar_puntos = st.checkbox("Mostrar puntos de datos", value=True)
                paletas_matplotlib = sorted(plt.colormaps())
                Ausenco_pallet = ['#101820','#004764',"#c6d1da",'#0095c8',"#b6c41d",
                  "#ffdb49",'#db7121','#a1292f','#662766',"#7E5B68",
                  "#267730","#4EBB5C","#4F5B66"]
                paletas_opciones = ['Ausenco'] + paletas_matplotlib

                paleta_seleccionada = st.selectbox("Selecciona una paleta de colores:", options=paletas_opciones, index=0)

                # Widget para seleccionar rango de zoom (en escala logarítmica)
                st.header("Configuración de Zoom")
                zoom_habilitado = st.checkbox("Activar vista con zoom")

                if zoom_habilitado:

                    x_min_zoom= st.number_input("Rango inferior", value=min_log, placeholder=min_log,step=0.001,format="%.3f")   
                    
                    x_max_zoom= st.number_input("Rango superior", value=max_log, placeholder=max_log,step=0.001,format="%.3f")  
                
                Agrupar_muestras = st.checkbox("Agrupar muestras") 

                    
            except Exception as e:
                st.error(f"Error al cargar el archivo: {e}")
        else:
            st.info("Por favor suba un archivo Excel")
            #st.stop()
        
    # Contenido principal
    if df_transformado is not None and df_plasticidad is not None:
        min_log = df_transformado['log_Tamaño'].min()
        max_log = df_transformado['log_Tamaño'].max()
        
        # Obtenemos las muestras únicas
        muestras_unicas = df_transformado['Muestra'].unique()
        
        # Creamos dos columnas para controles y gráfico
        col1, col2 = st.columns([1, 3])
        
        with col1:
            st.subheader("Título del gráfico")
            titulo_grafico = st.text_input(
                "Título del gráfico",
                value="Granulometría "
            )

            st.subheader("Selección de Curvas")
            
            # Selección múltiple de muestras a mostrar
            muestras_seleccionadas = st.multiselect(
                "Seleccione las muestras a graficar",
                options=muestras_unicas,
                default=muestras_unicas
            )
            
            # Selección de colores para cada muestra
            st.subheader("Personalización de Colores")
            colores = {}
            # Paleta de colores por defecto

            #palette = plt.colormaps.get_cmap('tab10').resampled(len(muestras_seleccionadas))
            # Obtener colores según selección
            colores_usar = fn["obtener_colores"](paleta_seleccionada, len(muestras_seleccionadas))#fn.obtener_colores(paleta_seleccionada, len(muestras_seleccionadas))

            palette = ListedColormap(colores_usar)
            option = st.selectbox(
                "Seleccionar la muestra para personalizar el color",
                options=muestras_seleccionadas)

            

            for i, muestra in enumerate(muestras_seleccionadas):
                color_default = palette(i)
                color_default_hex = '#%02x%02x%02x' % (
                    int(color_default[0]*255),
                    int(color_default[1]*255),
                    int(color_default[2]*255)
                )
                colores[muestra] = color_default_hex

            colores[option] = st.color_picker(
            f"Color para {option}",
            value=colores[option],
            )
        
            Grupo2 = None
            Nombre_grupos = None
            if Agrupar_muestras:
                st.subheader("Nombres de los grupos")
                Nombre_G1 = st.text_input(
                    "Grupo 1",
                    value="Grupo 1"
                )

                Nombre_G2 = st.text_input(
                    "Grupo 2",
                    value="Grupo 2"
                )

                Grupo2 = st.multiselect(
                    Nombre_G2,
                    muestras_seleccionadas,
                    default=None,
                )
                Nombre_grupos=[Nombre_G1,Nombre_G2]
        
        with col2:
            st.header("Gráfico Principal")
            
            # Crear y mostrar el gráfico principal
            fig_principal = fn["crear_grafico"](
                df_transformado, columna_tamaño, muestras_seleccionadas,
                colores, None, mostrar_puntos, zoom=False,
                titulo=titulo_grafico, Agrupar_muestras=Agrupar_muestras,
                grupo2=Grupo2, nombre_grupos=Nombre_grupos
            )
            # fn.crear_grafico(df_transformado, columna_tamaño, muestras_seleccionadas, colores, None, 
            #                             mostrar_puntos,zoom=False,titulo=titulo_grafico,
            #                             Agrupar_muestras=Agrupar_muestras,grupo2=Grupo2,nombre_grupos=Nombre_grupos)
            st.pyplot(fig_principal)
            
            # Botón para descargar el gráfico principal
            buf = BytesIO()
            fig_principal.savefig(buf, format="png", dpi=300, bbox_inches="tight")
            st.download_button(
                label="Descargar gráfico principal",
                data=buf.getvalue(),
                file_name=f'{titulo_grafico}_principal.png',
                mime="image/png"
            )
            
            # Mostrar gráfico con zoom si está habilitado
            if zoom_habilitado:
                st.header("Vista con Zoom")
                fig_zoom = fn["crear_grafico"](
                                df_transformado, columna_tamaño, muestras_seleccionadas,
                                colores, (x_min_zoom, x_max_zoom), mostrar_puntos,
                                zoom=zoom_habilitado, titulo=titulo_grafico,
                                Agrupar_muestras=Agrupar_muestras,
                                grupo2=Grupo2, nombre_grupos=Nombre_grupos
                            )
                # fn.crear_grafico(df_transformado, columna_tamaño, muestras_seleccionadas, colores, 
                #                     (x_min_zoom, x_max_zoom), mostrar_puntos,zoom=zoom_habilitado,titulo=titulo_grafico,
                #                     Agrupar_muestras=Agrupar_muestras,grupo2=Grupo2,nombre_grupos=Nombre_grupos)
                st.pyplot(fig_zoom)
                
                # Botón para descargar el gráfico con zoom
                buf_zoom = BytesIO()
                fig_zoom.savefig(buf_zoom, format="png", dpi=300, bbox_inches="tight")
                st.download_button(
                    label="Descargar gráfico con zoom",
                    data=buf_zoom.getvalue(),
                    file_name=f'{titulo_grafico}_detalle.png',
                    mime="image/png"
                )
            
            dfs = {}  # Diccionario para guardar los dataframes por muestra
            for muestra in muestras_seleccionadas:
                IP_muestra = df_plasticidad[(df_plasticidad['Muestra'] == muestra) & (df_plasticidad['Nombre'] == 'IP')]['Limite'].values[0]
                LL_muestra =  df_plasticidad[(df_plasticidad['Muestra'] == muestra) & (df_plasticidad['Nombre'] == 'LL')]['Limite'].values[0]
                
                Sample_processed = fn["texture_preprocess"](
                                        df=df_transformado, muestra=muestra, LL=LL_muestra, PI=IP_muestra
                                    )
                #fn.texture_preprocess(df=df_transformado,muestra=muestra,LL=LL_muestra, PI=IP_muestra)
                if Sample_processed.Finos is not None:
                    if  Sample_processed.Finos > 50:
                        fn["USCS_finos"](Sample_processed) # fn.USCS_finos(Sample_processed)
                    else:
                        fn["USCS_granular"](Sample_processed)#fn.USCS_granular(Sample_processed)
                else:
                    Sample_processed.Errores = '\n'.join(Sample_processed.Errores) if Sample_processed.Errores else None
                df_m = fn["to_dataframe"](Sample_processed)#fn.to_dataframe(Sample_processed)
                dfs[muestra] = df_m.rename(columns={df_m.columns[1]: muestra})

            # Convertir los dicts a lista y hacer merge secuencial por "Variable"
            df_final = reduce(lambda left, right: pd.merge(left, right, on='Variable', how='outer'), dfs.values())

            unidades = {
                'D10': 'D10 (mm)',
                'D30': 'D30 (mm)',
                'D60': 'D60 (mm)',
                'Arcillas': 'Arcillas (%)',
                'Arenas': 'Arenas (%)',
                'Gravas': 'Gravas (%)',
                'Finos': 'Finos (%)',
                'Limos': 'Limos (%)'
            }
            df_final['Variable'] = df_final['Variable'].replace(unidades)

            # Mostrar datos originales
            if st.checkbox("Mostrar datos originales",value=True):
                st.dataframe(df_final)
            # Botón para descargar los datos procesados
            buf_datos = BytesIO()
            df_final.to_excel(buf_datos, index=False, engine='openpyxl')

            # 2. Aplicar formato con openpyxl
            buf_datos.seek(0)
            wb = load_workbook(buf_datos)
            ws = wb.active

            # Definir formato
            negrita = Font(bold=True)
            borde = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = borde
                    if cell.column == 1:  # primera columna
                        cell.font = negrita

            # 3. Guardar de nuevo en buffer
            buf_datos = BytesIO()
            wb.save(buf_datos)
            buf_datos.seek(0)

            # 4. Botón de descarga en Streamlit
            st.download_button(
                label="Descargar datos procesados",
                data=buf_datos,
                file_name=f'{titulo_grafico}_Clasificación_USCS.xlsx',
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.warning("Por favor, cargue un archivo Excel para comenzar.")

with tabs[1]:
    st.header("Manual de Usuario 📘")
    
    st.markdown("""
    ### Cómo usar la aplicación:
    1. Cargue un archivo de datos granulométricos y de plasticidad en el formato indicado (Abajo se presenta un ejemplo del formato de tabla requerida la cual puede ser descargada).
    2. Seleccione las muestras a procesar.
    3. Personalice los colores de las curvas y el título del gráfico.
    4. Active la opción de zoom si desea enfocar un rango específico de tamaños.
    5. Agrupe las muestras si es necesario y defina el nombre de los grupos.   
    6. Visualice y Descargue los resultados de clasificación USCS.
    
    
    > ⚠️ **Notas importantes**:

    - Ante la falta de datos, dejar la plantilla sin valores en las celdas correspondientes.
    - Asegúrese de que los nombres de las columnas coincidan con el formato del modelo.
    - Ante falta de datos granulométricos exactos (Arcillas, Limos y Arenas), se realizarán **interpolaciones lineales** cuando sea posible.
    - La aplicación **no valida los datos**, por lo que es responsabilidad del usuario asegurarse del formato y contenido correcto.
    - Si faltan los datos de **LL y/o IP**, se asume **plasticidad baja** (bajo la línea "A") para suelos finos, y clasificación **ML u OL** para suelos gruesos.
    """)
    
    # Ejemplo de DataFrame modelo

    df_modelo = pd.DataFrame({
        'Nombre': ['LL', 'IP', 'Tamaño', 76.2, 63.5, 50.8, 38.1, 25.400000000000002,
       19.05, 12.700000000000001, 9.525, 4.75, 2, 1.651, 0.833, 0.6,
       0.425, 0.42, 0.3, 0.246, 0.212, 0.147, 0.104, 0.074, 0.06812,
       0.056229999999999995, 0.053, 0.04641, 0.043000000000000003, 0.038],
        'BH-01': [20, 6, '% Pasante', 100, 74.9009900990099, np.nan, 69.4539603960396,
       61.09653465346535, 57.67475247524752, 47.488613861386135,
       43.317821782178214, 36.413861386138606, 30.931683168316823, np.nan,
       np.nan, np.nan, 13.696534653465335, np.nan, np.nan, 12.09059405940593, np.nan,
       11.826732673267315, np.nan, 11.763366336633652, np.nan, np.nan,
       11.2549504950495, np.nan, 10.919801980198017, 10.669306930693068],
       'BH-02':[ np.nan, np.nan, '% Pasante', 100, 92.66483516483517,np.nan,
       92.66483516483517, 78.72417582417583, 70.86923076923077,
       60.52582417582418, 56.354395604395606, 47.72857142857143,
       38.925274725274726,np.nan,np.nan,np.nan, 15.95054945054946,np.nan,np.nan,
       13.280219780219795,np.nan, 12.176923076923089,np.nan,
       11.760439560439579,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan],
       'BH-03':[60, 40, '% Pasante', 100, 66.85140073081608,np.nan,
       48.745432399512794, 43.30085261875762, 40.657734470158346,
       36.12058465286236, 34.79293544457978, 31.997563946406828,
       28.440925700365412,np.nan,np.nan,np.nan, 16.24847746650427,np.nan,np.nan,
       15.444579780755177,np.nan, 15.310596833130333,np.nan, 15.2557856272838,
      np.nan,np.nan, 14.168392204628503,np.nan, 13.873672350791722,
       13.38581607795372],
       'BH-04':[50, 10, '% Pasante', 100, 100, 60.91475409836066,
       52.299672131147545, 35.24196721311476, 29.595737704918037,
       24.031475409836077, 22.713770491803288, 17.250819672131158,np.nan,
       14.129508196721318, 9.968852459016404, 7.695409836065579,
       5.587213114754107,np.nan, 3.7613114754098405,np.nan,
       2.6459016393442596, 1.9409836065573813, 1.4144262295081944,
       1.119016393442621,np.nan,np.nan, 0.9245901639344254,np.nan,
       0.8150819672131178, 0.7534426229508284],
       'BH-05':[np.nan,np.nan, '% Pasante', 100, 100, 53.96971214017522,
       49.38585732165207, 31.66070087609512, 28.154317897371712,
       22.731664580725905, 20.262828535669584, 15.163829787234036,np.nan,
       11.643554443053816, 7.223153942428027, 4.988485607008755,
       3.6036295369211473,np.nan, 2.6455569461827224,np.nan,
       2.0386733416770966, 1.6822277847309124, 1.4708385481852275,
       1.3239048811013703,np.nan,np.nan, 1.252065081351688,np.nan,
       1.2138923654568146, 1.1936170212765944],
       'BH-06':[30, 50, '% Pasante', 100, 100, 50.44140625, 45.7359375,
       38.89015625, 36.247499999999995, 33.0371875, 31.078437500000007,
       27.79468750000001,np.nan, 23.001875000000013, 16.02171875000002,
       12.579375000000013, 10.378437500000018,np.nan, 8.384218750000016,
      np.nan, 7.297187500000021, 6.289531250000024, 5.404843750000026,
       4.666250000000019,np.nan,np.nan, 4.3720312500000205,np.nan,
       4.223281250000014, 4.110156250000017]
    })

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_modelo.to_excel(writer, index=False, sheet_name='Ejemplo')

    processed_data = output.getvalue()

    # Mostrar tabla y botón
    st.subheader("Ejemplo de formato de datos requerido:")
    st.dataframe(df_modelo)

    st.download_button(
        label="📥 Descargar formato ejemplo (.xlsx)",
        data=processed_data,
        file_name='formato_ejemplo.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ) 
with tabs[2]:
    #st.header("Metodología")	
    st.markdown(   """
    ## 1. Introducción
    Esta aplicación permite realizar el análisis granulométrico de suelos, facilitando tanto la visualización de curvas granulométricas como la clasificación de las muestras. Los datos de entrada deben proporcionarse en formato Excel, con columnas que representen los tamaños de partícula y sus correspondientes valores para cada muestra.

    La clasificación de los suelos se realiza utilizando el **Sistema Unificado de Clasificación de Suelos (USCS)**, de acuerdo con lo establecido por la normativa **ASTM D2487-06** [(ver norma)](https://store.astm.org/d2487-06.html).
            
    ## 2. Determinación de contenidos de arcillas, limos y arenas:
                
    Para estimar el contenido de cada fracción textural (arcillas, limos, arenas y gravas), el sistema primero verifica si estos datos están presentes en el archivo cargado. En caso afirmativo, se utilizan directamente. Si uno o más valores están ausentes, se aplica **interpolación lineal en escala logarítmica**, siempre que existan datos mayores y menores al límite de fracción buscado.

    Los diámetros correspondientes a cada fracción textural siguen los criterios de la **Asociación Americana de Funcionarios de Transporte y Carreteras Estatales (AASHTO)**:

    | Fracción | Tamaño del grano (mm) |
    |----------|-----------------------|
    | Grava    | 76,2 - 2              |
    | Arenas   | 2 - 0,075             |
    | Limos    | 0,075 - 0,002         |
    | Arcillas | <0,002                |
                 
    ## 3. Determinación de D10, D30 y D60
    Los percentiles granulométricos D10, D30 y D60 se obtienen mediante **interpolación lineal en escala logarítmica**, a partir de los datos granulométricos disponibles para cada muestra.


    ## 4. Supuestos ante la falta de datos y sus limitaciones
    - Si faltan los valores de **Límite Líquido (LL)** y/o **Índice de Plasticidad (IP)**, se asume una **plasticidad baja** (por debajo de la línea "A") para suelos con más del 50% de finos, asignando una clasificación tentativa **ML** u **OL** según corresponda.
    - La aplicación **no valida ni corrige los datos de entrada**. Es responsabilidad del usuario asegurar que el archivo tenga el formato y contenido adecuado.
    - En ausencia de valores directos de contenido textural (arcillas, limos o arenas), se aplicará interpolación solo si existen datos suficientes para ello.
    - Se asume que los suelos **no son orgánicos**.

    ## 5. Errores comunes
    En los resultados de clasificación se adjuntan los errores encontrados durante el procesamiento de las muestras. Estos errores pueden incluir:
    - **Faltantes de datos** necesarios para la clasificación.
    - **Supuestos aplicados** debido a la ausencia de información clave.
    - **Fallas de interpolación**, cuando no se dispone de datos suficientes para estimar un valor (es decir, cuando faltan datos por encima y por debajo del punto buscado).
""")

# Definir ícono según el tema actual
active_theme = get_active_theme_key()
if active_theme in ["soft_dark", "dark", "Dark Midnight"]:
    icono_tema = ":material/light_mode:"  # Sol para cambiar a modo claro
else:
    icono_tema = ":material/dark_mode:"  # Luna para cambiar a modo oscuro

# Mostrar el botón con ícono dinámico
with st.sidebar:
    if st.button(f"{icono_tema}", key="toggle_theme"):
        st_theme_changer(
            themes_data=theme_data,
            render_mode="next",
            rerun_whole_st=True,
            key="theme_next"
        )
